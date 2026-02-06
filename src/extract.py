# src/extract.py
from __future__ import annotations

from pathlib import Path
import io

import pandas as pd


def get_raw_row_count(path: Path) -> int:
    """
    Fast-ish row count without parsing the full table into a DataFrame.

    CSV:
      - counts newline rows (minus header). Uses binary read for speed.

    XLS/XLSX:
      - uses openpyxl in read_only mode and returns max_row - 1 (header).
      - still has to read workbook metadata, but avoids full dataframe parsing.
    """
    suffix = path.suffix.lower()

    if suffix == ".csv":
        n = 0
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                n += chunk.count(b"\n")

        # If file doesn't end with newline, last line won't be counted by \n.
        try:
            with open(path, "rb") as f:
                f.seek(-1, io.SEEK_END)
                last = f.read(1)
            if last != b"\n":
                n += 1
        except Exception:
            pass

        # subtract header row if any rows exist
        return max(n - 1, 0)

    if suffix in {".xlsx", ".xls"}:
        try:
            from openpyxl import load_workbook  # type: ignore

            wb = load_workbook(filename=path, read_only=True, data_only=True)
            ws = wb.active
            return max(int(ws.max_row) - 1, 0)
        except Exception:
            df = pd.read_excel(path)
            return int(max(len(df), 0))

    # Unknown type: fallback
    try:
        df = pd.read_csv(path)
        return int(max(len(df), 0))
    except Exception:
        return 0


def read_table_clean_cols(path: Path) -> pd.DataFrame:
    """Read a CSV/XLSX file, normalize column names, and add source row position."""
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        df = pd.read_excel(path)
    else:
        df = pd.read_csv(path)

    df.columns = [str(c).strip() for c in df.columns]

    # Add a stable 1-based "Excel-like" row number INCLUDING header offset.
    # Header is row 1, so first data row is row 2.
    if "source_row_num" not in df.columns:
        df.insert(0, "source_row_num", (df.index.astype(int) + 2))

    return df


# Backwards-compat alias (older code imported this name)
read_csv_clean_cols = read_table_clean_cols
